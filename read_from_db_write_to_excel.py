from connection import connection
from openpyxl import Workbook
from openpyxl import load_workbook

conn_string = r'SERVER=localhost\SQLEXPRESS;DATABASE=demo;Trusted_Connection=yes;'

def read_data_from_db():
    query = " select * from city "
    # no params
    params = []
    conn = connection(conn_string=conn_string,query=query, params=params)
    city_data = conn.fetch_query()
    return city_data

def write_to_default_ws_row_col(cities):
    wb = Workbook(write_only=True)
    city_ws = wb.create_sheet("City")
    # write header
    city_ws.append(["Id", "City Name"])

    # write data
    for city in cities:
        city_id = city[0]
        city_name = city[1]
        city_ws.append([city_id, city_name])

    wb.save("default.xlsx")

def write_to_specific_cell(cities):
    # first create worksheet with write only mode
    wb = Workbook(write_only=True)
    city_ws = wb.create_sheet("City")
    wb.save("specific.xlsx")

    # then edit workbook with load_workbook
    wb2 = load_workbook("specific.xlsx")
    city_ws = wb2["City"]
    #  save header starting at cell B4
    city_ws.cell(row=4,column=2).value = "Id"
    city_ws.cell(row=4, column=3).value="City Name"

    # save data starting at cell B5
    for i,row in enumerate(cities):
        for j, col in enumerate(row):
            city_ws.cell(row=(5 + i), column=(2 + j)).value = col

    wb2.save("specific.xlsx")


if __name__ == "__main__":
    city_data = read_data_from_db()
    write_to_default_ws_row_col(cities= city_data)
    write_to_specific_cell(cities= city_data)
